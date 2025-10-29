import streamlit as st
import openpyxl
from io import BytesIO
import re

st.set_page_config(page_title="Excel Global Replace", page_icon="📗")

st.title("📗 Excel Global Replace Tool")

uploaded_file = st.file_uploader("📤 Upload Excel file (.xlsx or .xlsm)", type=["xlsx", "xlsm"])
find_text = st.text_input("🔍 Find this value:")
replace_text = st.text_input("✏️ Replace with:")
case_insensitive = st.checkbox("🔡 Case-insensitive search", value=False)

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file, data_only=False, keep_links=True)
    st.write("Sheets found:", wb.sheetnames)

    if st.button("🔄 Replace across all sheets"):
        count = 0
        log = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue

                    val_str = str(val)
                    find_str = str(find_text)
                    repl_str = str(replace_text)

                    if case_insensitive:
                        pattern = re.compile(re.escape(find_str), re.IGNORECASE)
                        new_val = pattern.sub(repl_str, val_str)
                    else:
                        new_val = val_str.replace(find_str, repl_str)

                    if new_val != val_str:
                        # try to restore type (int/float if possible)
                        try:
                            cell.value = float(new_val) if new_val.replace('.', '', 1).isdigit() else new_val
                        except:
                            cell.value = new_val
                        count += 1
                        if len(log) < 20:
                            log.append(f"{sheet.title}!{cell.coordinate}: {val} → {cell.value}")

        if count > 0:
            st.success(f"✅ Replaced {count} occurrence(s) of '{find_text}' with '{replace_text}' across all sheets.")
            st.code("\n".join(log))
        else:
            st.warning("⚠️ No matches found — check value format or case sensitivity.")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="⬇️ Download modified Excel file",
            data=output,
            file_name="modified.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
